import json
import openpyxl

# Categories for caption classification
CATEGORIES = [
     #Paste the exact same content you used for classifying the captions here.
 ]


# Load categorized captions from JSON file

categorized_captions_filename = "File name of your categorized captions goes here"
with open(categorized_captions_filename, "r", encoding="utf-8") as f:
    categorized_captions = json.load(f)

# Create a new Excel workbook
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = "Title of your worksheet goes here"

# Extract unique categories and create headers (starting from column A)
all_categories = sorted(set(category.strip()  # Strip any trailing whitespace
                               for entry in categorized_captions
                               for category in entry.split("]")[0][1:].split(",")))

for col_num, category in enumerate(all_categories, start=1):  # Start at column 1
    cell = worksheet.cell(row=1, column=col_num, value=category)
    cell.font = openpyxl.styles.Font(bold=True)  # Make header bold

# Track the next available row for each category
next_available_row = {category: 2 for category in all_categories}  # Start from row 2

# Process captions and place them in the correct columns
for entry in categorized_captions:
    categories_str, caption = entry.split("]", 1)
    categories = [cat.strip() for cat in categories_str[1:].split(",")]

    for category in categories:
        col_num = all_categories.index(category) + 1
        row_num = next_available_row[category]
        worksheet.cell(row=row_num, column=col_num, value=caption)
        next_available_row[category] += 1

# Save the workbook

workbook_name = "Name your excel sheet"
workbook.save(workbook_name)
